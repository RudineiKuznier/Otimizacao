
from enum import Enum
import xlwings as xw
import threading
import openpyxl


# Guia de uso para o rudinei :

# tab = Tabela(linha=i)
# parametros = tab.pegarParametros()
# parametros.muy
# parametros.mux1
# parametros.sigmax1
# parametros.sigmay
# parametros.mux2
# parametros.sigmax2
# parametros.linha_salvar
# tab.salvarColunaSaida(coluna=2,valor=-40028922,linha_lwb=i)
# tab.salvarColunaSaida(coluna=2,valor=-40028922,linha_divpad=i)

# Cria o mutex
mutex = threading.Lock()

# inicio da primeira linha coluna
LINHA_INICIO = 0
COLUNA_INICIO = 24

# diz a quantidade de colunas que separa uma tabela de outra
NUM_COLUNAS_ENTRE_TABELAS = 0

# Defines de tabelas
NUMLINHAS       = 3
NUMCOLUNAS      = 15
TABELA_MATLAB = "Stock_Data_in_days_cv_0,2_5V.xlsx"

TABELAS_PAGINA = "Matlab_Data"

LINHA_INICIO_RESPOSTA_LWB     =  43
LINHA_INICIO_RESPOSTA_DESVPAD =  48

class Parametros:
    def __init__ (self,muy : float,mux1 : float,sigmax1 : float,sigmay : float, mux2 : float, sigmax2 : float, linha_salvar : int, coluna_salvar : int,nome_tabela : str ,pagina_tabela : str):
        self.muy = muy
        self.mux1 = mux1
        self.sigmax1 = sigmax1
        self.sigmay = sigmay
        self.mux2 = mux2
        self.sigmax2 = sigmax2
        self.linha_salvar = linha_salvar
        self.coluna_salvar = coluna_salvar
        self.nome_tabela = nome_tabela
        self.pagina_tabela = pagina_tabela

    def __repr__(self):
        return (f"Parâmetros da tabela : \n muy = {self.muy} \n mux1 = {self.mux1} \n sigmax1 = {self.sigmax1} \n sigmay = {self.sigmay} \n mux2 = {self.mux2} \n sigmax2 = {self.sigmax2} \n linha = {self.linha_salvar} \n coluna = {self.coluna_salvar} \n ")


class Posicoes(Enum) :

    MUY             = (8,4) # iguais
    MUX1             = (18,4) # diferentes
    SIGMAX1          = (23,4) # diferentes
    SIGMAY          = (28,4) # iguais

    MUX2             = (33,4) 
    SIGMAX2          = (38,4)

    SAIDADESVPAD    = (50,4)
    SAIDALBW        = (45,4)

    @property
    def linha(self):
        return self.value[0]
    
    @property
    def coluna(self):
        return self.value[1]
    
    @property
    def par(self):
        return {self.value[0],self.value[1]}
    

class Tabela :
    '" TABELA É NÃO SINCRONIZADA (NÃO MANTÉM O ARQUIVO XLSX ABERTO), SE PRECISAR, CHAMA ATUALIZAR "'
    def __init__(self,pagina : str, localENome: str):
        self.tabelas = []
        mutex.acquire()
        try:
            workbook = openpyxl.load_workbook(localENome, data_only=True)
            print (f" --------- LENDO DE ---------\n Tabela. : {localENome} \n Página : {pagina} \n")
            # Verifica se a página existe
            if pagina in workbook.sheetnames:
                sheet = workbook[pagina]
            else:
                # Lista as páginas disponíveis se a desejada não existir
                print(f"Página {pagina} não encontrada.")
                workbook.close()
                exit(1) 
            for linha in range(NUMLINHAS):
                for coluna in range(NUMCOLUNAS):
                    muy = sheet.cell(row=Posicoes.MUY.linha + linha + LINHA_INICIO, column=Posicoes.MUY.coluna + coluna + COLUNA_INICIO).value or 0.0
                    mux1 = sheet.cell(row=Posicoes.MUX1.linha + linha + LINHA_INICIO, column=Posicoes.MUX1.coluna + coluna + COLUNA_INICIO).value or 0.0
                    sigmax1 = sheet.cell(row=Posicoes.SIGMAX1.linha + linha + LINHA_INICIO, column=Posicoes.SIGMAX1.coluna + coluna + COLUNA_INICIO).value or 0.0
                    sigmay = sheet.cell(row=Posicoes.SIGMAY.linha + linha + LINHA_INICIO, column=Posicoes.SIGMAY.coluna + coluna + COLUNA_INICIO).value or 0.0
                    mux2 = sheet.cell(row=Posicoes.MUX2.linha + linha + LINHA_INICIO, column=Posicoes.MUX2.coluna + coluna + COLUNA_INICIO).value or 0.0
                    sigmax2 = sheet.cell(row=Posicoes.SIGMAX2.linha + linha + LINHA_INICIO, column=Posicoes.SIGMAX2.coluna + coluna + COLUNA_INICIO).value or 0.0
                    
                    parametro = Parametros(muy=muy, mux1=mux1, sigmax1=sigmax1, sigmay=sigmay, mux2=mux2, sigmax2=sigmax2, linha_salvar=linha + LINHA_INICIO,coluna_salvar= coluna + COLUNA_INICIO,nome_tabela=localENome,pagina_tabela=pagina)

                    #print(parametro)

                    self.tabelas.append(parametro)
            
            workbook.close()
        except Exception as e:
            print(f"Erro ao carregar arquivo: {e}")
        finally:
            mutex.release()
        return
    
    def pegarParametros(self) -> list[Parametros] :
        return self.tabelas
    
    def salvarColunaSaida(self, valor: float, parametro : Parametros, linha_lwb: int = None, linha_desvpad: int = None):
        # Pede permissão
        mutex.acquire()
        try:
            try:
                # Carregar o arquivo
                workbook = openpyxl.load_workbook(parametro.nome_tabela)
                
                # Selecionar a página específica
                sheet = workbook[parametro.pagina_tabela]
                
                if (linha_desvpad != None):
                    linha = linha_desvpad + LINHA_INICIO_RESPOSTA_DESVPAD
                elif (linha_lwb != None):
                    linha = linha_lwb + LINHA_INICIO_RESPOSTA_LWB
                else:
                    print("ERRO AO TENTAR SALVAR NA TABELA, NÃO FOI PASSADO O PARÂMETRO DA LINHA DE SALVAMENTO.")
                    exit(1) 
                
                # Substituir o valor na posição (linha, coluna + 4)
                sheet.cell(row=parametro.linha_salvar + linha, column=parametro.coluna_salvar + 4).value = valor
                
                # Salvar o arquivo
                workbook.save(parametro.nome_tabela)
                workbook.close()
                
                # print(f"Valor {valor} salvo em {parametro.pagina_tabela}({ parametro.linha_salvar + linha}, {parametro.coluna_salvar + 4})")
                
            except FileNotFoundError:
                print(f"Erro: Arquivo '{parametro.nome_tabela}' não encontrado")
            except KeyError:
                print(f"Erro: Página '{parametro.pagina_tabela}' não encontrada. Páginas disponíveis: {workbook.sheetnames}")
            except PermissionError:
                print(f"Erro: Sem permissão para escrever no arquivo '{parametro.nome_tabela}'")
            except Exception as e:
                print(f"Erro ao salvar arquivo: {e}")

        finally:
            # Devolve o mutex
            mutex.release()
        
        return
        


# tab = Tabela(pagina= TABELAS_PAGINA,localENome= TABELA_MATLAB)
# parametros = tab.pegarParametros()
# for i in parametros :
#     tab.salvarColunaSaida(valor=-1,parametro=i,linha_lwb=0)