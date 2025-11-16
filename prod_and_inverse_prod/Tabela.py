import threading
import openpyxl
from Particoes_na_tabela import Particao, PosicoesParametros, SUPLIERS_POSICOES, FACTORS_POSICOES, DISTRIBUTORS_POSICOES, RETAILERS_POSICOES

mutex = threading.Lock()


class Parametros:
    def __init__(self, muy, mux1, sigmax1, sigmay, mux2, sigmax2,
                 linha_salvar, coluna_salvar, nome_tabela, pagina_tabela,
                 reorder, sigxart, sigyart, constart, muxart,
                 posicoes_parametros: PosicoesParametros):
        self.muy = muy
        self.mux1 = mux1
        self.sigmax1 = sigmax1
        self.sigmay = sigmay
        self.mux2 = mux2
        self.sigmax2 = sigmax2
        self.linha_salvar = linha_salvar
        self.coluna_salvar = coluna_salvar
        self.nome_tabela = nome_tabela
        self.pagina_tabela = pagina_tabela
        self.reorder = reorder
        self.sigxart = sigxart
        self.sigyart = sigyart
        self.constart = constart
        self.muxart = muxart
        self.lw = 0
        self.desv = 0
        self.prob = 0
        self.probart = 0
        self.posicoes_parametros = posicoes_parametros  # 🔹 chave para salvar corretamente

    def __repr__(self):
        return (f"Tabela : {self.nome_tabela} , Página : {self.pagina_tabela}"
                f"Parâmetros: muy={self.muy}, mux1={self.mux1}, sigmax1={self.sigmax1}, "
                f"sigmay={self.sigmay}, mux2={self.mux2}, sigmax2={self.sigmax2}, "
                f"linha salvar ={self.linha_salvar}, coluna salvar={self.coluna_salvar}")


class Tabela:
    def __init__(self):
        self.tabelas = []
        self.arquivo = None
        self.localENome = ""

    def carregarArquivoParaRam(self, localENome: str):
        self.arquivo = openpyxl.load_workbook(localENome, data_only=False)
        self.localENome = localENome

    def removerArquivoDaRam(self):
        self.arquivo.close()

    def carregarMatriz(self, particao: Particao, pagina: str, tabela: int) -> list[Parametros]:
        """Carrega uma matriz com base na definição de partição passada."""
        lista = []
        if self.arquivo is None:
            print("Erro ao tentar carregar matriz, arquivo não carregado previamente.")
            return []

        try:
            if pagina not in self.arquivo.sheetnames:
                raise ValueError(f"Página '{pagina}' não encontrada no arquivo.")
            sheet = self.arquivo[pagina]

            print(f" --------- LENDO MATRIZ {tabela} ---------")

            pos = particao.POSICOES_PARAMETROS
            shift_coluna = particao.COLUNA_INICIO + (tabela * (particao.PASSO_COLUNA + particao.QTD_COLUNAS))
            shift_linha = particao.LINHA_INICIO

            for linha in range(particao.QTD_LINHAS):
                for coluna in range(particao.QTD_COLUNAS):
                    sigxart = sheet.cell(row=pos.SIGMAXART + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    sigyart = sheet.cell(row=pos.SIGMAYART + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    constart = sheet.cell(row=pos.CONSTART + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    muxart = sheet.cell(row=pos.MUXART + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    muy = sheet.cell(row=pos.MUY + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    mux1 = sheet.cell(row=pos.MUX1 + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    sigmax1 = sheet.cell(row=pos.SIGMAX1 + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    sigmay = sheet.cell(row=pos.SIGMAY + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    mux2 = sheet.cell(row=pos.MUX2 + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    sigmax2 = sheet.cell(row=pos.SIGMAX2 + linha + shift_linha, column=shift_coluna + coluna).value or 0.0
                    reorder_c = sheet.cell(row=pos.REODER_C + linha + shift_linha, column=shift_coluna + coluna).value or 0.0

                    parametro = Parametros(
                        muy=muy, mux1=mux1, sigmax1=sigmax1, sigmay=sigmay,
                        mux2=mux2, sigmax2=sigmax2,
                        linha_salvar=linha,
                        coluna_salvar=coluna + shift_coluna,
                        nome_tabela=self.localENome, pagina_tabela=pagina,
                        reorder=reorder_c, sigxart=sigxart, sigyart=sigyart,
                        constart=constart, muxart=muxart,
                        posicoes_parametros=pos
                    )
                    if (False):
                        print(f"\nLendo células - linha_base={linha}, shift_linha={shift_linha}, coluna={coluna}, shift_coluna={shift_coluna}")
                        print(f"SIGMAXART: row={pos.SIGMAXART + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.SIGMAXART + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"SIGMAYART: row={pos.SIGMAYART + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.SIGMAYART + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"CONSTART: row={pos.CONSTART + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.CONSTART + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"MUXART: row={pos.MUXART + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.MUXART + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"MUY: row={pos.MUY + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.MUY + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"MUX1: row={pos.MUX1 + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.MUX1 + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"SIGMAX1: row={pos.SIGMAX1 + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.SIGMAX1 + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"SIGMAY: row={pos.SIGMAY + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.SIGMAY + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"MUX2: row={pos.MUX2 + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.MUX2 + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"SIGMAX2: row={pos.SIGMAX2 + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.SIGMAX2 + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"REODER_C: row={pos.REODER_C + linha + shift_linha}, column={shift_coluna + coluna}, valor={sheet.cell(row=pos.REODER_C + linha + shift_linha, column=shift_coluna + coluna).value}")
                        print(f"Valores finais: sigxart={sigxart}, sigyart={sigyart}, constart={constart}, muxart={muxart}, muy={muy}, mux1={mux1}, sigmax1={sigmax1}, sigmay={sigmay}, mux2={mux2}, sigmax2={sigmax2}, reorder_c={reorder_c}")
                        
                    lista.append(parametro)

            return lista

        except Exception as e:
            print(f"Erro ao carregar arquivo: {e}")
            return []

            
        except Exception as e:
            print(f"Erro ao carregar arquivo: {e}")

    def salvarEmLote(self, parametros: list[Parametros]):
        """Salva as colunas de saída (desvio, lw, prob, probart) com base na partição salva em cada parâmetro."""
        if (self.arquivo != None):
        
            if not parametros:
                print("Nenhum parâmetro para salvar.")
                return

            pos = getattr(parametros[0], "posicoes_parametros", None)
            if pos is None:
                print("Erro: as posições de parâmetros não foram definidas para esta partição.")
                return

            try:
                if parametros[0].pagina_tabela not in self.arquivo.sheetnames:
                    print(f"Erro: Página '{parametros[0].pagina_tabela}' não encontrada. Páginas: {self.arquivo.sheetnames}")
                    self.arquivo.close()
                    return

                for p in parametros:
                    pos = p.posicoes_parametros
                    self.arquivo[parametros[0].pagina_tabela].cell(row=p.linha_salvar + pos.SAIDADESVPAD, column=p.coluna_salvar).value = p.desv
                    self.arquivo[parametros[0].pagina_tabela].cell(row=p.linha_salvar + pos.SAIDALBW, column=p.coluna_salvar).value = p.lw
                    self.arquivo[parametros[0].pagina_tabela].cell(row=p.linha_salvar + pos.SAIDAPROB, column=p.coluna_salvar).value = p.prob
                    self.arquivo[parametros[0].pagina_tabela].cell(row=p.linha_salvar + pos.SAIDAPROBART, column=p.coluna_salvar).value = p.probart

                self.arquivo.save(parametros[0].nome_tabela)
            except Exception as e:
                print(f"Erro ao salvar arquivo: {e}")
        else :
            print("Nenhum arquivo aberto, não foi possível salvar.")

# ----------------- TESTE -----------------
# if __name__ == "__main__":
#     tabela = Tabela()
#     tabela.carregarArquivoParaRam(localENome="Stock_Data_in_days_cv_02_5V.xlsx")
#     supliers = tabela.carregarMatriz(SUPLIERS_POSICOES,"Main_variables",tabela=0)
#     factors = tabela.carregarMatriz(FACTORS_POSICOES,"Main_variables",tabela=0)
#     distributors = tabela.carregarMatriz(DISTRIBUTORS_POSICOES,"Main_variables",tabela=0)
#     retails = tabela.carregarMatriz(RETAILERS_POSICOES,"Main_variables",tabela=0)

#     parametros = supliers + factors + distributors + retails
#     for p in parametros:
#         p.lw = -20
#         p.desv = -21
#         p.prob = -22
#         p.probart = -25
        
#     tabela.salvarEmLote(parametros)


#     tabela.removerArquivoDaRam()
