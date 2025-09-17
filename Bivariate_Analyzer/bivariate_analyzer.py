import numpy as np
from scipy.stats import multivariate_normal
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

class BivariateCDFAnalyzer:
    def __init__(self, mu, sigma, x1_range, x2_range, num_points):
        self.mu = np.array(mu)
        self.sigma = np.array(sigma)
        self.x1_min, self.x1_max = x1_range # (min, max) grade X1
        self.x2_min, self.x2_max = x2_range # (min, max) grade X2
        self.num_points = num_points #dimensao grade

        self.x1_grid = None
        self.x2_grid = None
        self.X1_grid = None
        self.X2_grid = None
        self.X = None
        self.p = None
        self.product_grid = None
        self.products_flat = None
        self.Z = None # Plotagem 3D

        self._initialize_grid_and_data()

    def _initialize_grid_and_data(self):

        self.x1_grid = np.linspace(self.x1_min, self.x1_max, self.num_points)
        self.x2_grid = np.linspace(self.x2_min, self.x2_max, self.num_points)

        self.X1_grid, self.X2_grid = np.meshgrid(self.x1_grid, self.x2_grid)
        self.X = np.column_stack([self.X1_grid.ravel(), self.X2_grid.ravel()])

        # Calcular a CDF
        dist = multivariate_normal(mean=self.mu, cov=self.sigma)
        self.p = dist.cdf(self.X)

        # Calcular a matriz de produtos (X1 * X2)
        self.product_grid = self.X1_grid * self.X2_grid
        self.products_flat = self.product_grid.ravel()

        # Reshape da CDF para a plotagem 3D
        self.Z = self.p.reshape(self.num_points, self.num_points)

    def export_to_excel(self, file_name='Probability distributions.xlsx', sheet_name='Sheet3'):

        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Cabeçalhos
        ws['D4'] = 'CDF (p)'
        ws['E4'] = 'X1'
        ws['F4'] = 'X2'
        ws['G4'] = 'X1 * X2'
        ws['M4'] = ''

        # CDF
        # ws[f'D{5 + i}'] = float(f"{val:.7f}") #aumentar a precisão
        for i, val in enumerate(self.p):
            ws[f'D{5 + i}'] = val

        # X1
        for i, val in enumerate(self.X[:, 0]):
            ws[f'E{5 + i}'] = val

        # X2
        for i, val in enumerate(self.X[:, 1]):
            ws[f'F{5 + i}'] = val

        # Matriz de produtos (X1 * X2)
        for i, val in enumerate(self.products_flat):
            ws[f'G{5 + i}'] = val

        # Escrever a matriz Z
        start_row_z = 5
        start_col_z = 13 # 13ª coluna

        for r_idx in range(self.Z.shape[0]):
            for c_idx in range(self.Z.shape[1]):
                ws.cell(row=start_row_z + r_idx, column=start_col_z + c_idx, value=self.Z[r_idx, c_idx])

        wb.save(file_name)
        print(f"Dados exportados com sucesso para '{file_name}' na planilha '{sheet_name}'.")

    def find_closest_cdf_values(self, input_value, num_results):
        valid_indices = np.where(self.products_flat <= input_value)[0]

        if not valid_indices.size:
            return []

        filtered_products = self.products_flat[valid_indices]
        filtered_cdf_values = self.p[valid_indices]
        filtered_X_coords = self.X[valid_indices]

        #differences = np.abs(self.products_flat - input_value)
        #sorted_indices = np.argsort(differences)

        differences = input_value - filtered_products

        sorted_indices = np.argsort(differences)

        top_indices = sorted_indices[:num_results]

        results = []
        for idx in top_indices:
            results.append({
                "cdf_value": filtered_cdf_values[idx],
                "product_X1X2": filtered_products[idx],
                "X1": filtered_X_coords[idx, 0],
                "X2": filtered_X_coords[idx, 1]
            })
        return results

    def plot_cdf_3d(self, title='CDF da Distribuição Normal Bivariada', zlabel='CDF (p)'):
        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, projection='3d')

        surf = ax.plot_surface(self.X1_grid, self.X2_grid, self.Z, cmap='viridis')

        ax.set_title(title)
        ax.set_xlabel('X1')
        ax.set_ylabel('X2')
        ax.set_zlabel(zlabel)

        fig.colorbar(surf, shrink=0.5, aspect=5, label=zlabel)
        plt.show()