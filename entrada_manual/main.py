import os
import json
import re
import pandas as pd
from openpyxl import load_workbook


def excel_col_to_index(col):
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord('A') + 1)
    return idx - 1

def parse_range(cell_range):
    start, end = cell_range.split(":")
    col_start = re.match(r"[A-Z]+", start).group()
    row_start = int(re.search(r"\d+", start).group())
    col_end = re.match(r"[A-Z]+", end).group()
    row_end = int(re.search(r"\d+", end).group())
    return (
        excel_col_to_index(col_start),
        row_start - 1,
        excel_col_to_index(col_end),
        row_end - 1
    )

def read_excel_values_only(path, sheet_name):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows():
        data.append([cell.value for cell in row])
    return pd.DataFrame(data)

def rotate_table(df, spec, ordem):
    c0, r0, c1, r1 = parse_range(spec["range"])
    rows = []

    for i, r in enumerate(range(r0, r1 + 1), start=1):
        for j, c in enumerate(range(c0, c1 + 1), start=1):
            val = df.iat[r, c]
            if val is None:
                val = 0

            rows.append({
                "Grupo": spec["grupo"],          # interno
                "Tabela numero": spec["tabela"],
                "Linha": i,
                "Coluna": j,
                "Variavel": spec["variavel"],
                "Valor": val,
                "_ordem_variavel": ordem
            })

    return pd.DataFrame(rows)

def process_planilha(
    input_excel,
    tables_txt,
    sheet_name="Main_Variables",
    output_excel=None
):
    if output_excel is None:
        base, ext = os.path.splitext(input_excel)
        output_excel = f"{base}_saida{ext}"

    # Ler planilha
    df_raw = read_excel_values_only(input_excel, sheet_name)

    # Ler especificação
    with open(tables_txt, "r", encoding="utf-8") as f:
        tables = json.load(f)

    # Ordem FIXA das variáveis
    ordem_desejada = [
        "P(Reorder)",
        "P(Estado da arte)",
        "Reorder",
        "Var(Var(d))",
        "Var(Var(Cap))",
        "CV",
        "Planned Stock",
        "Variance coeff - z(S2nd)",
        "demand",
        "Média(E(RLT))",
        "Cap",
        "Var(d-cap)",
        "Desvpad(Var(S2nd))",
        "DesvPad(Var(E(RLT)))",
        "Média(d-Cap)",
        "DesvPad(Var(d-cap))",
        "LB",
        "DesvPad(ND*ELT)",
        "DesvPad(Cap)",
        "z(RLT)",
        "Aleatória S2nd",
        "Radicand",
        "SQRT",
        "ND*ELT",
        "Aleatória RLT",
        "ActualRotw",
        "vMaxNumOfCyclesw",
        "DesvPad(demand)",
        "DesvPad(E(RLT))",
        "DesvPad(D*E(RLT))",
        "Média(D*E(RLT))",
        "Reorder(Estado da arte)",
        "Erro > 0,01%",
        "DesvPad(D*ELT)"
    ]


    MAPEAMENTO_GRUPO = {
        "Suppliers": {
            "linha_offset": 0,     # 1,2,3
            "coluna_offset": 0     # 1..15
        },
        "Factories": {
            "linha_offset": 3,     # 1→4, 2→5, 3→6
            "coluna_offset": 15    # 1→16, 2→17, 3→18
        },
        "Distributors": {
            "linha_offset": 6,     # 1→7, 2→8
            "coluna_offset": 18    # 1→19, 2→20, 3→21
        },
        "Retailers": {
            "linha_offset": 8,     # 1→9, ...
            "coluna_offset": 21    # 1→22, 2→23, 3→24
        }
    }


    dfs = []
    for ordem, spec in enumerate(tables):
        dfs.append(rotate_table(df_raw, spec, ordem))

    df_long = pd.concat(dfs, ignore_index=True)

    df_long = df_long.drop_duplicates(
        subset=["Grupo", "Tabela numero", "Linha", "Coluna", "Variavel"],
        keep="first"
    )

    resultado_por_grupo = {}

    for grupo in sorted(df_long["Grupo"].unique()):
        df_grupo = df_long[df_long["Grupo"] == grupo]
        tabelas_pivotadas = []

        for tabela_num in sorted(df_grupo["Tabela numero"].unique()):
            df_tab = df_grupo[df_grupo["Tabela numero"] == tabela_num]

            df_pivot = df_tab.pivot_table(
                index=["Linha", "Coluna"],
                columns="Variavel",
                values="Valor",
                aggfunc="first",
                dropna=False
            ).reset_index()

            # Garantir todas as variáveis
            for var in ordem_desejada:
                if var not in df_pivot.columns:
                    df_pivot[var] = 0

            # 🔑 Aplicar mapeamento conceitual
            if grupo in MAPEAMENTO_GRUPO:
                df_pivot["Linha"]  = df_pivot["Linha"]  + MAPEAMENTO_GRUPO[grupo]["linha_offset"]
                df_pivot["Coluna"] = df_pivot["Coluna"] + MAPEAMENTO_GRUPO[grupo]["coluna_offset"]

            # Inserir identificadores (SEM Grupo)
            df_pivot.insert(0, "Tabela numero", tabela_num)

            # Ordem final
            df_pivot = df_pivot[
                ["Tabela numero", "Linha", "Coluna"] + ordem_desejada
            ]

            tabelas_pivotadas.append(df_pivot)

        resultado_por_grupo[grupo] = pd.concat(
            tabelas_pivotadas, ignore_index=True
        )


    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        for grupo, df_final in resultado_por_grupo.items():
            df_final = df_final.sort_values(
                ["Tabela numero", "Linha", "Coluna"]
            )

            df_final.to_excel(
                writer,
                sheet_name=grupo,
                index=False
            )

            print(f"✓ {grupo}: {len(df_final)} linhas salvas")

    print(f"\n✓ Arquivo salvo: {output_excel}")


if __name__ == "__main__":
    process_planilha(
        input_excel="Stock_Data_in_days_cv_02_3Var_Rev03.xlsx",
        #input_excel="Stock_Data_in_days_cv_1_3Var_rev04.xlsx",
        #input_excel="Stock_Data_in_days_cv_06_3Var_Rev04.xlsx",
        tables_txt="entrada_completa.txt",
        sheet_name="Main_variables"
    )
